# -*- coding: utf-8 -*-
"""
送货单 Excel 解析模块
- 每个 sheet 对应一张送货单（如 SG-260813-0001）
- 从送货单中提取表头信息（送货单号、送货日期）与明细行
"""
import re
from datetime import datetime

import openpyxl

# 明细表的关键列名
DETAIL_HEADER = [
    "产品名称", "材料名称", "颜色", "备注", "包装",
    "产品规格", "数量", "单位", "包装明细",
]


def _find_detail_header(ws):
    """定位明细表头所在行。返回表头行号（1-based）或 None。"""
    for row_idx in range(1, min(ws.max_row + 1, 15)):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 10)]
        text = "".join(str(v) for v in row_vals if v)
        # 表头行应包含"产品名称"和"包装明细"
        if "产品名称" in text and "包装明细" in text:
            return row_idx
    return None


def _find_field(ws, key, col=7):
    """在某个列中查找 'key: value' 形式的字段值。"""
    for row_idx in range(1, 6):
        val = ws.cell(row=row_idx, column=col).value
        if val and str(key) in str(val):
            # 提取冒号后的内容
            m = re.search(rf"{key}[：:]\s*(.*)", str(val))
            if m:
                return m.group(1).strip()
    return ""


def extract_main_no(delivery_no):
    """
    从完整送货单号提取订单主号（去掉末尾分号）。
    例：SG-260813-0001 -> SG-260813
    """
    s = str(delivery_no or "").strip()
    m = re.match(r"^(.*?)-\d{2,4}$", s)
    return m.group(1) if m else s


def parse_delivery_note(ws, customer_default=""):
    """
    解析单个 sheet（一张送货单分单）。
    返回 dict: {delivery_no, main_no, date, customer, items:[...]}
    """
    delivery_no = _find_field(ws, "送货单号")
    date_str = _find_field(ws, "送货日期")
    main_no = extract_main_no(delivery_no)

    header_row = _find_detail_header(ws)
    if header_row is None:
        return None

    # 读取表头，建立列索引
    header = {}
    for c in range(1, 10):
        name = ws.cell(row=header_row, column=c).value
        if name:
            header[str(name).strip()] = c

    items = []
    for r in range(header_row + 1, ws.max_row + 1):
        first = ws.cell(row=r, column=header.get("产品名称", 1)).value
        if first is None or str(first).strip() == "":
            continue
        row = {}
        for key, col in header.items():
            row[key] = ws.cell(row=r, column=col).value
        # 只保留真正的明细行：数量列必须为数字
        qty = row.get("数量")
        if qty is None:
            continue
        if isinstance(qty, str) and not qty.strip().replace(",", "").isdigit():
            continue
        # 跳过合计行
        if str(first).strip().startswith("合计"):
            continue
        items.append(row)

    return {
        "delivery_no": delivery_no,
        "main_no": main_no,
        "date": date_str,
        "customer": customer_default,
        "items": items,
    }


def parse_workbook(path, customer_default=""):
    """
    解析整本送货单 Excel，返回 [parse_delivery_note] 列表。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    notes = []
    for ws in wb.worksheets:
        note = parse_delivery_note(ws, customer_default)
        if note and note["items"]:
            notes.append(note)
    return notes


def normalize_date(date_str):
    """将日期字符串统一为 YYYY-MM-DD，用于照片文件夹匹配。"""
    if not date_str:
        return ""
    s = str(date_str).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s
